# crm.py
import os
import json
import uuid
from typing import Optional
from io import BytesIO
from datetime import datetime

from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import RedirectResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from db import (
	init_db,
	get_flows, create_flow, delete_flow, move_flow,
	get_blocks, get_block, create_block, update_block, delete_block,
	next_position, swap_positions,
	get_stats, get_users,
	get_flow_triggers, set_flow_trigger, delete_flow_trigger,

	# ✅ flow modes (off/manual/auto)
	get_flow_modes, set_flow_mode,

	# ✅ flow actions (after flow -> start flow)
	get_flow_actions, upsert_flow_action, delete_flow_action,

	# ✅ broadcasts (new)
	list_broadcasts, create_broadcast, delete_broadcast, set_broadcast_active,
)

from seed import seed as run_seed  # ✅ автосид

app = FastAPI()
templates = Jinja2Templates(directory="templates")

os.makedirs("media", exist_ok=True)
app.mount("/media", StaticFiles(directory="media"), name="media")


@app.on_event("startup")
async def startup():
	await init_db()

	# ✅ автосид только если flows пустые (первый запуск на новой БД)
	try:
		flows = await get_flows()
		if not flows:
			await run_seed()
	except Exception:
		pass


# ─────────────────────────────────────────────────────────────
# Helpers

def _unit_to_seconds(unit: str) -> int:
	u = (unit or "").strip().lower()
	if u == "seconds":
		return 1
	if u == "minutes":
		return 60
	if u == "hours":
		return 3600
	return 86400  # days default


def _seconds_to_value_unit(total_seconds: int, preferred_unit: str = "minutes") -> tuple[int, str]:
	"""
	Конвертируем секунды -> (value, unit) для удобного UI.
	Теперь поддерживаем seconds/minutes/hours/days.
	"""
	s = int(total_seconds or 0)
	p = (preferred_unit or "minutes").strip().lower()
	if p not in ("seconds", "minutes", "hours", "days"):
		p = "minutes"

	if s <= 0:
		return 0, p

	# предпочитаем “красивые” деления
	if s % 86400 == 0:
		return s // 86400, "days"
	if s % 3600 == 0:
		return s // 3600, "hours"
	if s % 60 == 0:
		return s // 60, "minutes"

	# иначе — показываем в секундах (честнее, чем округлять вверх)
	return s, "seconds"


def _value_unit_to_seconds(value: int, unit: str) -> int:
	v = int(value or 0)
	if v < 0:
		v = 0
	return v * _unit_to_seconds(unit)


def _safe_filename(name: str) -> str:
	n = (name or "").strip()
	n = os.path.basename(n)
	n = n.replace("\x00", "").replace("\n", " ").replace("\r", " ").strip()
	return n


def _norm_mode(mode: str) -> str:
	m = (mode or "").strip().lower()
	if m not in ("off", "manual", "auto"):
		return "off"
	return m


# broadcasts helpers
def _norm_schedule_type(t: str) -> str:
	tt = (t or "").strip().lower()
	return tt if tt in ("monthly", "interval_days") else "monthly"


def _norm_days_of_month(s: str) -> str:
	ss = (s or "").strip()
	return ss or "1"


def _clamp_int(v: int, lo: int, hi: int) -> int:
	try:
		vv = int(v)
	except Exception:
		vv = lo
	return max(lo, min(hi, vv))


# ─────────────────────────────────────────────────────────────
# INDEX (FLOWS + STATS + USERS + TRIGGERS + MODES + ACTIONS + BROADCASTS)

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
	flows = await get_flows()
	stats = await get_stats()
	users = await get_users(200)

	# ✅ flow modes
	try:
		modes = await get_flow_modes()
	except Exception:
		modes = {}

	# triggers (offset + enabled)
	raw = await get_flow_triggers()
	triggers_map = {}

	for f in flows:
		triggers_map[f] = {
			"flow": f,
			"enabled": False,
			"offset_value": 0,
			"offset_unit": "days",
			"offset_seconds": 0,
			"mode": _norm_mode(modes.get(f, "off")),
		}

	for t in raw:
		flow = (t.get("flow") or "").strip()
		if not flow:
			continue

		enabled = int(t.get("is_active", 0) or 0)
		offset_seconds = int(t.get("offset_seconds", 0) or 0)
		val, unit = _seconds_to_value_unit(offset_seconds, preferred_unit="days")

		if flow not in triggers_map:
			triggers_map[flow] = {
				"flow": flow,
				"enabled": bool(enabled),
				"offset_value": int(val),
				"offset_unit": unit,
				"offset_seconds": offset_seconds,
				"mode": _norm_mode(modes.get(flow, "off")),
			}
		else:
			triggers_map[flow].update({
				"enabled": bool(enabled),
				"offset_value": int(val),
				"offset_unit": unit,
				"offset_seconds": offset_seconds,
				"mode": _norm_mode(modes.get(flow, triggers_map[flow].get("mode", "off"))),
			})

	# ✅ сценарии "после flow"
	try:
		actions = await get_flow_actions(None)
	except Exception:
		actions = []

	for a in actions:
		val, unit = _seconds_to_value_unit(int(a.get("delay_seconds", 0) or 0), preferred_unit="minutes")
		a["delay_value"] = int(val)
		a["delay_unit"] = unit

	# ✅ broadcasts list
	try:
		broadcasts = await list_broadcasts()
	except Exception:
		broadcasts = []

	for b in broadcasts:
		b["is_all_users"] = (b.get("target_user_id") is None)
		b["schedule_type"] = _norm_schedule_type(b.get("schedule_type", "monthly"))
		b["at_hour"] = int(b.get("at_hour", 12) or 12)
		b["at_minute"] = int(b.get("at_minute", 0) or 0)

	return templates.TemplateResponse(
		"index.html",
		{
			"request": request,
			"flows": flows,
			"stats": stats,
			"users": users,
			"triggers": triggers_map,
			"actions": actions,
			"broadcasts": broadcasts,
		},
	)


# ─────────────────────────────────────────────────────────────
# FLOW MODE + TRIGGERS routes

@app.post("/flow/{flow}/trigger")
async def flow_trigger_save(
	flow: str,
	mode: str = Form("off"),
	offset_value: int = Form(0),
	offset_unit: str = Form("days"),
):
	flow = (flow or "").strip()
	if not flow:
		return RedirectResponse("/", status_code=302)

	mode = _norm_mode(mode)
	await set_flow_mode(flow, mode)

	offset_value = int(offset_value or 0)
	if offset_value < 0:
		offset_value = 0

	unit = (offset_unit or "days").strip().lower()
	if unit not in ("seconds", "minutes", "hours", "days"):
		unit = "days"

	seconds = offset_value * _unit_to_seconds(unit)
	is_active = 1 if mode == "auto" else 0

	await set_flow_trigger(
		flow=flow,
		offset_seconds=int(seconds),
		is_active=int(is_active),
		trigger="after_start",
	)
	return RedirectResponse("/", status_code=302)


@app.post("/flow/{flow}/trigger/delete")
async def flow_trigger_delete(flow: str):
	flow = (flow or "").strip()
	if flow:
		await delete_flow_trigger(flow)
		try:
			await set_flow_mode(flow, "off")
		except Exception:
			pass
	return RedirectResponse("/", status_code=302)


# ─────────────────────────────────────────────────────────────
# FLOW ACTIONS

@app.post("/flow/action/upsert")
async def flow_action_upsert(
	after_flow: str = Form(""),
	target_flow: str = Form(""),
	is_active: int = Form(1),
	delay_value: int = Form(0),
	delay_unit: str = Form("minutes"),
	delay_seconds: int = Form(0),
):
	after_flow = (after_flow or "").strip()
	target_flow = (target_flow or "").strip()
	if not after_flow or not target_flow:
		return RedirectResponse("/", status_code=302)

	delay_unit = (delay_unit or "minutes").strip().lower()
	if delay_unit not in ("seconds", "minutes", "hours", "days"):
		delay_unit = "minutes"

	sec_from_ui = _value_unit_to_seconds(delay_value, delay_unit)
	delay = int(sec_from_ui if int(delay_value or 0) > 0 else int(delay_seconds or 0))
	if delay < 0:
		delay = 0

	await upsert_flow_action(
		after_flow=after_flow,
		target_flow=target_flow,
		delay_seconds=delay,
		is_active=1 if int(is_active) else 0,
		action_type="start_flow",
	)
	return RedirectResponse("/", status_code=302)


@app.post("/flow/action/{action_id}/delete")
async def flow_action_delete(action_id: int):
	try:
		await delete_flow_action(int(action_id))
	except Exception:
		pass
	return RedirectResponse("/", status_code=302)


# ─────────────────────────────────────────────────────────────
# BROADCASTS

@app.post("/broadcast/new")
async def broadcast_new(
	title: str = Form(""),
	flow: str = Form(""),
	target_mode: str = Form("all"),
	target_user_id: int = Form(0),
	schedule_type: str = Form("monthly"),
	days_of_month: str = Form("1"),
	interval_days: int = Form(30),
	at_hour: int = Form(12),
	at_minute: int = Form(0),
	is_active: int = Form(1),
):
	flow = (flow or "").strip()
	if not flow:
		return RedirectResponse("/", status_code=302)

	title = (title or "").strip() or f"Broadcast: {flow}"

	target_mode = (target_mode or "all").strip().lower()
	if target_mode not in ("all", "user"):
		target_mode = "all"

	tuid: Optional[int] = None
	if target_mode == "user":
		try:
			tuid = int(target_user_id)
		except Exception:
			tuid = None
		if not tuid or tuid <= 0:
			return RedirectResponse("/", status_code=302)

	schedule_type = _norm_schedule_type(schedule_type)
	days_of_month = _norm_days_of_month(days_of_month)
	interval_days = int(interval_days or 30)
	if interval_days < 1:
		interval_days = 1

	at_hour = _clamp_int(at_hour, 0, 23)
	at_minute = _clamp_int(at_minute, 0, 59)

	await create_broadcast(
		title=title,
		flow=flow,
		target_user_id=tuid,
		schedule_type=schedule_type,
		interval_days=interval_days,
		days_of_month=days_of_month,
		at_hour=at_hour,
		at_minute=at_minute,
		is_active=1 if int(is_active) else 0,
	)

	return RedirectResponse("/", status_code=302)


@app.post("/broadcast/{broadcast_id}/delete")
async def broadcast_delete(broadcast_id: int):
	try:
		await delete_broadcast(int(broadcast_id))
	except Exception:
		pass
	return RedirectResponse("/", status_code=302)


@app.post("/broadcast/{broadcast_id}/toggle")
async def broadcast_toggle(broadcast_id: int, is_active: int = Form(1)):
	try:
		await set_broadcast_active(int(broadcast_id), 1 if int(is_active) else 0)
	except Exception:
		pass
	return RedirectResponse("/", status_code=302)


# ─────────────────────────────────────────────────────────────
# EXPORT (XLSX)

@app.get("/export/users.xlsx")
async def export_users_xlsx():
	users = await get_users(50000)

	wb = Workbook()
	ws = wb.active
	ws.title = "bot_users"

	headers = [
		"user_id",
		"username",
		"first_seen_ts",
		"last_seen_ts",
		"first_seen_utc",
		"last_seen_utc",
		"starts_count",
		"messages_count",
	]
	ws.append(headers)

	def ts_to_utc_str(ts: int) -> str:
		if not ts:
			return ""
		return datetime.utcfromtimestamp(int(ts)).strftime("%Y-%m-%d %H:%M:%S")

	for u in users:
		ws.append([
			u.get("user_id"),
			u.get("username", ""),
			u.get("first_seen_ts"),
			u.get("last_seen_ts"),
			ts_to_utc_str(u.get("first_seen_ts")),
			ts_to_utc_str(u.get("last_seen_ts")),
			u.get("starts_count", 0),
			u.get("messages_count", 0),
		])

	for col_idx, h in enumerate(headers, start=1):
		ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(h) + 2)

	buf = BytesIO()
	wb.save(buf)
	buf.seek(0)

	filename = f"bot_users_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
	return StreamingResponse(
		buf,
		media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		headers={"Content-Disposition": f'attachment; filename="{filename}"'},
	)


# ─────────────────────────────────────────────────────────────
# FLOWS

@app.get("/flow/new")
async def flow_new_get():
	return RedirectResponse("/", status_code=302)


@app.post("/flow/new")
async def flow_new_post(name: str = Form("")):
	name = (name or "").strip()
	if not name:
		return RedirectResponse("/", status_code=302)

	await create_flow(name)
	try:
		await set_flow_mode(name, "off")
	except Exception:
		pass
	return RedirectResponse("/", status_code=302)


@app.post("/flow/{flow}/delete")
async def flow_delete(flow: str):
	await delete_flow(flow)
	return RedirectResponse("/", status_code=302)


@app.post("/flow/{flow}/up")
async def flow_up(flow: str):
	await move_flow(flow, "up")
	return RedirectResponse("/", status_code=302)


@app.post("/flow/{flow}/down")
async def flow_down(flow: str):
	await move_flow(flow, "down")
	return RedirectResponse("/", status_code=302)


@app.get("/flow/{flow}", response_class=HTMLResponse)
async def flow_page(request: Request, flow: str):
	blocks = await get_blocks(flow)
	return templates.TemplateResponse(
		"flow.html",
		{"request": request, "flow": flow, "blocks": blocks},
	)


# ─────────────────────────────────────────────────────────────
# BLOCKS

@app.get("/block/new", response_class=HTMLResponse)
async def new_block_page(request: Request, flow: str):
	flow = (flow or "").strip()
	if not flow:
		return RedirectResponse("/", status_code=302)

	await create_flow(flow)
	pos = await next_position(flow)

	flows = await get_flows()

	empty = {
		"id": 0,
		"flow": flow,
		"position": pos,
		"type": "text",
		"title": "",
		"text": "",
		"circle": "",
		"video": "",
		"buttons": "",
		"is_active": 1,

		# ✅ delay UI (seconds поддерживается корректно)
		"delay_value": 0,
		"delay_unit": "seconds",
		"delay": 1.0,

		"file_path": "",
		"file_kind": "",
		"file_name": "",

		"btn1_text": "",
		"btn1_url": "",
		"btn2_text": "",
		"btn2_url": "",
		"btn3_text": "",
		"btn3_url": "",
		"buttons_json": "",

		"gate_next_flow": "",
		"gate_button_text": "✅ Готов к следующему уроку",
		"gate_prompt_text": "👇 Нажми кнопку, чтобы перейти дальше",
		"gate_reminder_value": 0,
		"gate_reminder_unit": "hours",
		"gate_reminder_text": "",
	}

	return templates.TemplateResponse(
		"edit.html",
		{"request": request, "block": empty, "is_new": True, "flows": flows},
	)


@app.get("/block/{block_id}/edit", response_class=HTMLResponse)
async def edit_block_page(request: Request, block_id: int):
	block = await get_block(block_id)
	if not block:
		return RedirectResponse("/", status_code=302)

	flows = await get_flows()

	# buttons parse
	btns = []
	try:
		if block.get("buttons"):
			btns = json.loads(block["buttons"])
	except Exception:
		pass

	for i in range(3):
		block[f"btn{i+1}_text"] = ""
		block[f"btn{i+1}_url"] = ""

	for i, b in enumerate(btns[:3]):
		if isinstance(b, dict):
			block[f"btn{i+1}_text"] = b.get("text", "")
			block[f"btn{i+1}_url"] = b.get("url", "")

	block["buttons_json"] = block.get("buttons", "")

	# ✅ delay seconds -> value/unit
	delay_sec = int(float(block.get("delay", 1.0) or 0))
	dv, du = _seconds_to_value_unit(delay_sec, preferred_unit="minutes")
	block["delay_value"] = int(dv)
	block["delay_unit"] = du

	# gate reminder UI
	rem_sec = int(block.get("gate_reminder_seconds") or 0)
	val, unit = _seconds_to_value_unit(rem_sec, preferred_unit="hours")
	block["gate_reminder_value"] = int(val)
	block["gate_reminder_unit"] = unit

	if not (block.get("gate_button_text") or "").strip():
		block["gate_button_text"] = "✅ Готов к следующему уроку"
	if not (block.get("gate_prompt_text") or "").strip():
		block["gate_prompt_text"] = "👇 Нажми кнопку, чтобы перейти дальше"

	return templates.TemplateResponse(
		"edit.html",
		{"request": request, "block": block, "is_new": False, "flows": flows},
	)


@app.post("/block/save")
async def save_block(
	request: Request,
	block_id: int = Form(0),
	flow: str = Form(...),
	position: int = Form(...),
	type: str = Form(...),

	title: str = Form(""),
	text: str = Form(""),
	circle_path: str = Form(""),
	video_url: str = Form(""),
	is_active: int = Form(1),

	# delay
	delay_seconds: float = Form(1.0),
	delay_value: int = Form(0),
	delay_unit: str = Form("minutes"),

	file_path: str = Form(""),
	file_kind: str = Form(""),
	file_name: str = Form(""),

	btn1_text: str = Form(""),
	btn1_url: str = Form(""),
	btn2_text: str = Form(""),
	btn2_url: str = Form(""),
	btn3_text: str = Form(""),
	btn3_url: str = Form(""),
	buttons_json: str = Form(""),

	circle_file: UploadFile | None = File(None),
	attach_file: UploadFile | None = File(None),

	gate_next_flow: str = Form(""),
	gate_button_text: str = Form(""),
	gate_prompt_text: str = Form(""),
	gate_reminder_value: int = Form(0),
	gate_reminder_unit: str = Form("hours"),
	gate_reminder_text: str = Form(""),
):
	flow = (flow or "").strip()
	if not flow:
		return RedirectResponse("/", status_code=302)

	await create_flow(flow)

	# ✅ delay normalize (value/unit приоритетнее)
	du = (delay_unit or "minutes").strip().lower()
	if du not in ("seconds", "minutes", "hours", "days"):
		du = "minutes"

	if int(delay_value or 0) > 0:
		delay_final = float(_value_unit_to_seconds(delay_value, du))
	else:
		delay_final = float(delay_seconds or 0)

	if delay_final < 0:
		delay_final = 0.0

	# upload circle
	if circle_file and circle_file.filename:
		ext = os.path.splitext(circle_file.filename)[1].lower() or ".mp4"
		fname = f"{uuid.uuid4().hex}{ext}"
		with open(os.path.join("media", fname), "wb") as f:
			f.write(await circle_file.read())
		circle_path = f"/media/{fname}"

	# upload attachment
	if attach_file and attach_file.filename:
		orig_name = _safe_filename(attach_file.filename)
		ext = os.path.splitext(orig_name)[1].lower()
		fname = f"{uuid.uuid4().hex}{ext}" if ext else f"{uuid.uuid4().hex}"

		with open(os.path.join("media", fname), "wb") as f:
			f.write(await attach_file.read())

		file_path = f"/media/{fname}"
		file_name = orig_name

		ct = (attach_file.content_type or "").lower()
		if ct.startswith("image/"):
			file_kind = "photo"
		elif ct.startswith("video/"):
			file_kind = "video"
		elif ct.startswith("audio/"):
			file_kind = "audio"
		else:
			file_kind = "document"

	# buttons
	buttons = []
	for t, u in [(btn1_text, btn1_url), (btn2_text, btn2_url), (btn3_text, btn3_url)]:
		t = (t or "").strip()
		u = (u or "").strip()
		if t and u:
			buttons.append({"text": t, "url": u})

	buttons_final = ""
	if (buttons_json or "").strip():
		buttons_final = buttons_json.strip()
	elif buttons:
		buttons_final = json.dumps(buttons, ensure_ascii=False)

	# gate normalize
	gate_next_flow = (gate_next_flow or "").strip()
	gate_button_text = (gate_button_text or "").strip()
	gate_prompt_text = (gate_prompt_text or "").strip()

	gate_reminder_value = int(gate_reminder_value or 0)
	if gate_reminder_value < 0:
		gate_reminder_value = 0

	gate_reminder_unit = (gate_reminder_unit or "hours").strip().lower()
	if gate_reminder_unit not in ("seconds", "minutes", "hours", "days"):
		gate_reminder_unit = "hours"

	gate_reminder_seconds = gate_reminder_value * _unit_to_seconds(gate_reminder_unit)
	gate_reminder_text = (gate_reminder_text or "").strip()

	data = {
		"flow": flow,
		"position": int(position),
		"type": type,
		"title": title,
		"text": text,
		"circle": circle_path,
		"video": video_url,
		"buttons": buttons_final,
		"is_active": int(is_active),

		"delay": float(delay_final),

		"file_path": (file_path or "").strip(),
		"file_kind": (file_kind or "").strip(),
		"file_name": (file_name or "").strip(),

		"gate_next_flow": gate_next_flow,
		"gate_button_text": gate_button_text,
		"gate_prompt_text": gate_prompt_text,
		"gate_reminder_seconds": int(gate_reminder_seconds),
		"gate_reminder_text": gate_reminder_text,
	}

	if int(block_id) == 0:
		await create_block(data)
	else:
		await update_block(int(block_id), data)

	return RedirectResponse(f"/flow/{flow}", status_code=302)


@app.post("/block/{block_id}/delete")
async def delete_block_action(block_id: int, flow: str = Form(...)):
	await delete_block(block_id)
	return RedirectResponse(f"/flow/{flow}", status_code=302)


@app.post("/block/{block_id}/up")
async def move_up(block_id: int, flow: str = Form(...)):
	blocks = await get_blocks(flow)
	idx = next((i for i, b in enumerate(blocks) if b["id"] == block_id), None)
	if idx is not None and idx > 0:
		await swap_positions(blocks[idx]["id"], blocks[idx - 1]["id"])
	return RedirectResponse(f"/flow/{flow}", status_code=302)


@app.post("/block/{block_id}/down")
async def move_down(block_id: int, flow: str = Form(...)):
	blocks = await get_blocks(flow)
	idx = next((i for i, b in enumerate(blocks) if b["id"] == block_id), None)
	if idx is not None and idx < len(blocks) - 1:
		await swap_positions(blocks[idx]["id"], blocks[idx + 1]["id"])
	return RedirectResponse(f"/flow/{flow}", status_code=302)